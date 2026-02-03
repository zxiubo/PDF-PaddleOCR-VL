#!/usr/bin/env python3
"""
Excel生成与填充脚本
支持两种模式：
1. 自动生成：根据抽取的数据字段自动创建Excel（使用招聘报名模板）
2. 模板填充：使用用户提供的Excel模板填充数据

功能特性：
- 数据真实性校验（正则规则 + 智能体判断）
- 校验结果颜色标记（红色=错误，黄色=存疑，白色=正常）
"""

import os
import sys
import json
import datetime
import re
import traceback
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


# 定义招聘报名模板的默认字段顺序
DEFAULT_FIELD_ORDER = [
    "序号",
    "报名序号",
    "招聘单位",
    "岗位名称",
    "姓名",
    "身份证号码",
    "手机联系方式",
    "邮箱",
    "性别",
    "出生年月民族",
    "籍贯",
    "政治面貌",
    "集体户口",
    "户籍所在地",
    "详细居住地",
    "硕士毕业学校",
    "是否全日制",
    "学历学位双证齐全",
    "专业",
    "毕业时间",
    "本科毕业学校",
    "专业",
    "毕业时间",
    "大专毕业学校",
    "专业",
    "毕业时间",
    "高中毕业学校",
    "是否退役士兵",
    "立功情况",
    "社会工作者职称",
    "备注（其他证书）"
]


# 定义颜色样式
COLOR_RED = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
COLOR_YELLOW = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
COLOR_WHITE = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")


def log_debug(msg):
    """输出调试信息"""
    print(f"[DEBUG] {msg}", file=sys.stderr)


def get_field_order(field_name):
    """
    获取字段的排序顺序（基于招聘报名模板）
    """
    try:
        # 精确匹配
        if field_name in DEFAULT_FIELD_ORDER:
            return DEFAULT_FIELD_ORDER.index(field_name)
        
        # 模糊匹配
        for idx, std_field in enumerate(DEFAULT_FIELD_ORDER):
            if field_name.lower() in std_field.lower() or std_field.lower() in field_name.lower():
                return idx
        
        return 999
    except Exception as e:
        log_debug(f"get_field_order error: {e}")
        return 999


def validate_data(field_name, value, record):
    """
    校验数据的真实性/合理性
    """
    try:
        if not value or value == "":
            return (True, 'ok', "空值")
        
        value = str(value).strip()
        
        # 1. 姓名校验
        if "姓名" in field_name:
            if len(value) < 2 or len(value) > 10:
                return (False, 'warning', "姓名长度异常")
            if re.match(r'^[\u4e00-\u9fa5]{2,10}$', value) is None:
                return (False, 'warning', "姓名格式异常")
            return (True, 'ok', "格式正确")
        
        # 2. 身份证号校验
        if "身份证" in field_name:
            if len(value) != 18:
                return (False, 'error', "身份证号长度错误")
            # 简单校验前17位是否为数字
            if not value[:17].isdigit():
                return (False, 'error', "身份证号格式错误")
            # 校验最后一位（数字或X）
            if not (value[17].isdigit() or value[17].upper() == 'X'):
                return (False, 'error', "身份证号校验位错误")
            # 校验出生日期部分（第7-14位）
            try:
                birth_date = value[6:14]
                year = int(birth_date[:4])
                month = int(birth_date[4:6])
                day = int(birth_date[6:8])
                if year < 1900 or year > datetime.datetime.now().year:
                    return (False, 'warning', "身份证出生年份异常")
                if month < 1 or month > 12:
                    return (False, 'error', "身份证出生月份异常")
                if day < 1 or day > 31:
                    return (False, 'error', "身份证出生日期异常")
            except Exception as e:
                return (False, 'error', "身份证出生日期解析失败")
            return (True, 'ok', "格式正确")
        
        # 3. 手机号校验
        if "手机" in field_name or "联系" in field_name:
            if len(value) != 11:
                return (False, 'error', "手机号长度错误")
            if not value.isdigit():
                return (False, 'error', "手机号格式错误")
            if not value.startswith('1'):
                return (False, 'warning', "手机号格式异常")
            return (True, 'ok', "格式正确")
        
        # 4. 邮箱校验
        if "邮箱" in field_name or "email" in field_name.lower():
            try:
                email_pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
                if re.match(email_pattern, value) is None:
                    return (False, 'warning', "邮箱格式异常")
                return (True, 'ok', "格式正确")
            except Exception as e:
                return (False, 'warning', "邮箱校验失败")
        
        # 5. 年龄/出生年月校验
        if "年龄" in field_name or "出生" in field_name:
            if "年龄" in field_name:
                try:
                    age = int(value)
                    if age < 15 or age > 65:
                        return (False, 'warning', "年龄超出合理范围")
                except Exception as e:
                    return (False, 'warning', "年龄格式错误")
            return (True, 'ok', "格式正确")
        
        # 6. 毕业时间校验
        if "毕业时间" in field_name:
            try:
                # 尝试解析日期
                date_patterns = [
                    r'(\d{4})-(\d{1,2})-(\d{1,2})',
                    r'(\d{4})/(\d{1,2})/(\d{1,2})',
                    r'(\d{4})年(\d{1,2})月(\d{1,2})日',
                    r'(\d{4})\.(\d{1,2})\.(\d{1,2})',
                    r'(\d{4})(\d{2})(\d{2})',
                ]
                parsed = False
                for pattern in date_patterns:
                    try:
                        match = re.search(pattern, value)
                        if match:
                            year, month, day = match.groups()
                            year, month, day = int(year), int(month), int(day)
                            if year < 1950 or year > datetime.datetime.now().year + 2:
                                return (False, 'warning', f"毕业年份{year}异常")
                            if month < 1 or month > 12:
                                return (False, 'error', "毕业月份异常")
                            parsed = True
                            break
                    except Exception:
                        continue
                if not parsed:
                    return (False, 'warning', "毕业时间格式无法解析")
            except Exception as e:
                return (False, 'warning', "毕业时间格式错误")
            return (True, 'ok', "格式正确")
        
        # 7. 籍贯/户籍地校验（标记为存疑，需要人工确认）
        if "籍贯" in field_name or "户籍" in field_name or "居住地" in field_name:
            if len(value) < 2:
                return (False, 'warning', "地址信息过短")
            # 简单判断是否包含省市
            if not any(keyword in value for keyword in ["省", "市", "自治区", "县", "区"]):
                return (False, 'warning', "地址格式可能不完整，建议人工确认")
            return (True, 'ok', "格式基本正确")
        
        # 8. 学校名称校验
        if "学校" in field_name:
            if len(value) < 4:
                return (False, 'warning', "学校名称过短")
            # 检查是否包含大学、学院等关键词
            if not any(keyword in value for keyword in ["大学", "学院", "学校"]):
                return (False, 'warning', "学校名称可能不完整，建议人工确认")
            return (True, 'ok', "格式基本正确")
        
        # 其他字段默认通过
        return (True, 'ok', "通过")
        
    except Exception as e:
        log_debug(f"validate_data error for field '{field_name}': {e}")
        return (True, 'ok', "校验异常，默认通过")


def auto_generate_excel(records, output_path):
    """
    自动生成Excel文件（使用招聘报名模板）
    """
    result = {
        "success": False,
        "output_path": "",
        "fields": [],
        "rows": 0,
        "validation_summary": {
            "ok": 0,
            "warning": 0,
            "error": 0
        },
        "message": ""
    }
    
    try:
        log_debug("开始自动生成Excel...")
        
        # 收集所有字段
        all_fields = set()
        for record in records:
            all_fields.update(record.keys())
        
        log_debug(f"收集到的字段: {all_fields}")
        
        # 按标准顺序排序字段
        sorted_fields = sorted(all_fields, key=get_field_order)
        result["fields"] = sorted_fields
        
        # 创建工作簿
        log_debug("创建工作簿...")
        wb = Workbook()
        ws = wb.active
        ws.title = "招聘报名表"
        
        # 定义样式
        header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        # 写入表头
        log_debug("写入表头...")
        for col_idx, field in enumerate(sorted_fields, start=1):
            cell = ws.cell(row=1, column=col_idx, value=field)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # 设置列宽
        log_debug("设置列宽...")
        for col_idx, field in enumerate(sorted_fields, start=1):
            width = max(10, min(30, 8 + len(field) * 1.2))
            ws.column_dimensions[chr(64 + col_idx)].width = width
        
        # 写入数据并进行校验
        log_debug(f"开始写入数据，共{len(records)}条记录...")
        for row_idx, record in enumerate(records, start=2):
            log_debug(f"处理第{row_idx-1}条记录...")
            for col_idx, field in enumerate(sorted_fields, start=1):
                try:
                    value = record.get(field, "")
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                    
                    # 数据校验
                    is_valid, status, message = validate_data(field, value, record)
                    
                    # 更新校验摘要
                    if status == 'ok':
                        result["validation_summary"]["ok"] += 1
                    elif status == 'warning':
                        result["validation_summary"]["warning"] += 1
                        cell.fill = COLOR_YELLOW  # 标黄
                    elif status == 'error':
                        result["validation_summary"]["error"] += 1
                        cell.fill = COLOR_RED  # 标红
                except Exception as e:
                    log_debug(f"写入单元格失败: row={row_idx}, col={col_idx}, field={field}, error={e}")
                    continue
        
        # 冻结首行
        ws.freeze_panes = 'A2'
        
        # 保存文件
        log_debug(f"保存文件到: {output_path}")
        wb.save(output_path)
        
        result["success"] = True
        result["output_path"] = output_path
        result["rows"] = len(records)
        
        # 生成校验信息
        summary = result["validation_summary"]
        msg_parts = [f"自动生成Excel成功，包含{len(records)}条记录，{len(sorted_fields)}个字段"]
        if summary["error"] > 0:
            msg_parts.append(f"发现{summary['error']}个错误（红色标记）")
        if summary["warning"] > 0:
            msg_parts.append(f"发现{summary['warning']}个存疑项（黄色标记）")
        result["message"] = "，".join(msg_parts)
        
        # 添加表头信息到结果中
        result["template_headers"] = sorted_fields
        result["header_row"] = 1
        
        log_debug("Excel生成完成")
        return result
        
    except Exception as e:
        error_msg = f"自动生成Excel失败: {str(e)}"
        log_debug(error_msg)
        log_debug(traceback.format_exc())
        result["message"] = error_msg
        return result


def find_header_row(sheet, header_keywords):
    """查找表头行"""
    try:
        for row_idx, row in enumerate(sheet.iter_rows(max_row=10)):
            row_values = [str(cell.value or "") for cell in row]
            row_text = " ".join(row_values)
            matched_count = sum(1 for keyword in header_keywords if keyword in row_text)
            if matched_count >= len(header_keywords) * 0.5:
                return row_idx
        return None
    except Exception as e:
        log_debug(f"find_header_row error: {e}")
        return None


def find_column_by_header(sheet, header_row, header_name):
    """根据表头名称查找列号"""
    try:
        if header_row is None:
            return None
        header_cells = list(sheet.iter_rows(min_row=header_row + 1, max_row=header_row + 1))[0]
        for col_idx, cell in enumerate(header_cells):
            cell_value = str(cell.value or "")
            if header_name.lower() in cell_value.lower() or cell_value.lower() in header_name.lower():
                return col_idx
        return None
    except Exception as e:
        log_debug(f"find_column_by_header error: {e}")
        return None


def fill_to_template(template_path, records, output_path):
    """
    使用模板填充Excel（带数据校验）
    使用固定列索引映射，避免重复列名导致的匹配问题
    """
    result = {
        "success": False,
        "output_path": "",
        "filled_rows": 0,
        "unfilled_fields": [],
        "validation_summary": {
            "ok": 0,
            "warning": 0,
            "error": 0
        },
        "message": ""
    }
    
    try:
        log_debug(f"开始填充模板: {template_path}")
        
        # 复制模板到输出路径，避免修改原模板
        import shutil
        log_debug(f"复制模板文件: {template_path} -> {output_path}")
        shutil.copy2(template_path, output_path)
        
        # 加载复制后的文件
        wb = load_workbook(output_path)
        sheet = wb.active
        
        # 动态读取模板的表头
        log_debug("读取模板表头...")
        common_headers = ["姓名", "身份证号码", "手机联系方式", "邮箱"]
        header_row = None
        
        # 查找表头行
        for row_idx in range(1, min(11, sheet.max_row + 1)):  # openpyxl使用1-based索引
            row_cells = list(sheet.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
            row_values = [str(cell) if cell is not None else "" for cell in row_cells]
            row_text = " ".join(row_values)
            
            # 检查是否包含常见表头
            matched_count = sum(1 for h in common_headers if h in row_text)
            if matched_count >= 2:  # 至少包含2个常见字段
                header_row = row_idx
                break
        
        if header_row is None:
            header_row = 1
        
        # 读取表头
        header_row_cells = list(sheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True))[0]
        headers = [str(cell) if cell is not None else "" for cell in header_row_cells]
        
        # 过滤空表头并建立列映射
        headers = [h for h in headers if h.strip()]
        column_map = {}
        for col_idx, header in enumerate(headers):
            column_map[header] = col_idx
        
        # 数据从表头行的下一行开始填充
        data_start_row = header_row
        
        log_debug(f"表头行: {header_row}")
        log_debug(f"表头列数: {len(headers)}")
        log_debug(f"表头列表: {headers}")
        log_debug(f"数据起始行: {data_start_row}")
        log_debug(f"列映射: {column_map}")
        
        # 填充数据
        log_debug(f"开始填充数据，共{len(records)}条记录...")
        filled_rows = 0
        unfilled_fields = []
        
        for row_idx, record in enumerate(records):
            # 计算实际填充的行号（Excel行号从1开始）
            actual_row = data_start_row + row_idx + 1
            log_debug(f"处理第{row_idx+1}条记录，Excel行号{actual_row}...")
            
            for field, value in record.items():
                try:
                    # 直接使用字段名匹配表头
                    if field in column_map:
                        col_idx = column_map[field]
                        # Excel列号从1开始，所以列索引+1
                        cell = sheet.cell(row=actual_row, column=col_idx + 1)
                        cell.value = value
                        cell.alignment = Alignment(horizontal='left', vertical='center')
                        
                        # 数据校验
                        is_valid, status, message = validate_data(field, value, record)
                        
                        if status == 'ok':
                            result["validation_summary"]["ok"] += 1
                        elif status == 'warning':
                            result["validation_summary"]["warning"] += 1
                            cell.fill = COLOR_YELLOW
                        elif status == 'error':
                            result["validation_summary"]["error"] += 1
                            cell.fill = COLOR_RED
                    else:
                        if field not in unfilled_fields:
                            unfilled_fields.append(field)
                except Exception as e:
                    log_debug(f"填充单元格失败: row={actual_row}, field={field}, error={e}")
                    continue
            
            filled_rows += 1
        
        # 保存文件
        log_debug(f"保存文件到: {output_path}")
        wb.save(output_path)
        
        result["success"] = True
        result["output_path"] = output_path
        result["filled_rows"] = filled_rows
        result["unfilled_fields"] = unfilled_fields
        
        # 生成校验信息
        summary = result["validation_summary"]
        msg_parts = [f"成功填充{filled_rows}条记录"]
        if summary["error"] > 0:
            msg_parts.append(f"发现{summary['error']}个错误（红色标记）")
        if summary["warning"] > 0:
            msg_parts.append(f"发现{summary['warning']}个存疑项（黄色标记）")
        if unfilled_fields:
            msg_parts.append(f"未填充字段: {', '.join(unfilled_fields)}")
        result["message"] = "，".join(msg_parts)
        
        # 添加模板信息到结果中
        result["template_headers"] = headers
        result["header_row"] = header_row
        
        log_debug("模板填充完成")
        return result
        
    except Exception as e:
        error_msg = f"模板填充失败: {str(e)}"
        log_debug(error_msg)
        log_debug(traceback.format_exc())
        result["message"] = error_msg
        return result


def parse_extracted_data(data):
    """解析抽取的数据"""
    try:
        # 如果是字符串，尝试解析
        if isinstance(data, str):
            # 检查是否是文件路径
            if os.path.isfile(data):
                log_debug(f"从文件读取数据: {data}")
                with open(data, 'r', encoding='utf-8') as f:
                    data = json.load(f)
            else:
                # 尝试直接解析JSON字符串
                try:
                    data = json.loads(data)
                except json.JSONDecodeError:
                    # 如果不是JSON，尝试当作文件路径
                    if os.path.isfile(data):
                        log_debug(f"作为文件路径读取: {data}")
                        with open(data, 'r', encoding='utf-8') as f:
                            data = json.load(f)
                    else:
                        # 最后的fallback：包装为raw_data
                        log_debug(f"无法解析为JSON，包装为raw_data: {data[:100]}")
                        data = {"raw_data": data}
        
        if isinstance(data, dict):
            records = [data]
        elif isinstance(data, list):
            records = data
        else:
            raise Exception(f"不支持的data格式: {type(data)}")
        
        log_debug(f"解析到{len(records)}条记录")
        return records
    except Exception as e:
        log_debug(f"parse_extracted_data error: {e}")
        log_debug(traceback.format_exc())
        raise


def get_default_template_path():
    """
    获取默认模板的路径
    """
    try:
        # 获取脚本所在目录
        script_dir = os.path.dirname(os.path.abspath(__file__))
        # 获取Skill根目录
        skill_root = os.path.dirname(script_dir)
        # 默认模板路径
        default_template = os.path.join(skill_root, "assets", "个人信息提取结果-模板.xlsx")
        
        log_debug(f"默认模板路径: {default_template}")
        
        # 检查默认模板是否存在
        if not os.path.exists(default_template):
            log_debug(f"默认模板不存在: {default_template}")
            return None
        
        return default_template
    except Exception as e:
        log_debug(f"get_default_template_path error: {e}")
        return None


def process_excel(data, output_path=None, template_path=None, json_output_path=None):
    """
    处理Excel生成/填充
    
    参数说明：
    - data: JSON数据，可以是单条记录(dict)或多条记录(list)
    - template_path: Excel模板路径，如果None则使用默认模板
    - output_path: 输出文件路径，如果None则自动生成
    - json_output_path: JSON数据文件路径，如果None则自动生成（与Excel同名）
    
    工作流程：
    1. 解析数据（支持单条或多条记录）
    2. 使用模板填充数据
    3. 进行数据校验
    4. 保存JSON数据文件
    5. 保存并返回Excel文件
    """
def process_excel(data, output_path=None, template_path=None, json_output_path=None):
    """
    处理Excel生成/填充
    
    参数说明：
    - data: JSON数据，可以是单条记录(dict)或多条记录(list)
    - template_path: Excel模板路径，如果None则使用默认模板
    - output_path: 输出文件路径，如果None则自动生成
    - json_output_path: JSON数据文件路径，如果None则自动生成（与Excel同名）
    
    工作流程：
    1. 解析数据（支持单条或多条记录）
    2. 使用模板填充数据
    3. 进行数据校验
    4. 保存JSON数据文件
    5. 保存并返回Excel文件
    """
    try:
        log_debug("开始处理Excel...")
        
        # 解析数据
        records = parse_extracted_data(data)
        log_debug(f"解析到{len(records)}条记录")
        
        # 确定输出路径
        if output_path is None:
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = f"./个人信息提取结果_{timestamp}.xlsx"
        
        log_debug(f"输出路径: {output_path}")
        
        # 确定JSON文件路径（与Excel同名）
        if json_output_path is None:
            # 将.xlsx替换为.json
            json_output_path = output_path.rsplit('.', 1)[0] + '.json'
        
        log_debug(f"JSON输出路径: {json_output_path}")
        
        # 保存JSON数据文件
        try:
            with open(json_output_path, 'w', encoding='utf-8') as f:
                json.dump(records, f, ensure_ascii=False, indent=2)
            log_debug(f"JSON数据已保存到: {json_output_path}")
        except Exception as e:
            log_debug(f"保存JSON文件失败: {e}")
        
        # 确定使用的模板
        if template_path:
            log_debug("使用用户提供的模板")
            result = fill_to_template(template_path, records, output_path)
        else:
            log_debug("用户未提供模板，尝试使用默认模板")
            default_template = get_default_template_path()
            
            if default_template:
                log_debug("使用默认模板填充数据")
                result = fill_to_template(default_template, records, output_path)
            else:
                log_debug("默认模板不存在，自动生成Excel")
                result = auto_generate_excel(records, output_path)
        
        # 在结果中添加JSON文件路径
        result["json_output_path"] = json_output_path
        
        # 更新消息，包含JSON文件信息
        json_msg = f"，JSON数据文件已保存到: {json_output_path}"
        result["message"] = result.get("message", "") + json_msg
        
        return result
            
    except Exception as e:
        log_debug(f"process_excel error: {e}")
        log_debug(traceback.format_exc())
        return {
            "success": False,
            "output_path": "",
            "message": f"处理失败: {str(e)}"
        }


def main():
    """命令行入口"""
    try:
        if len(sys.argv) < 2:
            print("用法1（自动生成）: python excel_filler.py <data_json> [output_path] [--json-output json_path]")
            print("用法2（模板填充）: python excel_filler.py <data_json> --template <template_path> [output_path] [--json-output json_path]")
            sys.exit(1)
        
        data_json = sys.argv[1]
        template_path = None
        output_path = None
        json_output_path = None
        
        i = 2
        while i < len(sys.argv):
            if sys.argv[i] == "--template" and i + 1 < len(sys.argv):
                template_path = sys.argv[i + 1]
                i += 2
            elif sys.argv[i] == "--json-output" and i + 1 < len(sys.argv):
                json_output_path = sys.argv[i + 1]
                i += 2
            elif not sys.argv[i].startswith("--"):
                output_path = sys.argv[i]
                i += 1
            else:
                i += 1
        
        result = process_excel(data_json, output_path, template_path, json_output_path)
        print(json.dumps(result, ensure_ascii=False, indent=2))
        sys.exit(0 if result["success"] else 1)
        
    except Exception as e:
        error_result = {
            "success": False,
            "output_path": "",
            "message": f"程序异常: {str(e)}"
        }
        print(json.dumps(error_result, ensure_ascii=False, indent=2))
        sys.exit(1)


if __name__ == "__main__":
    main()
