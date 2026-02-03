#!/usr/bin/env python3
"""
读取Excel模板的表头
"""

import os
import sys
import json
import traceback
from openpyxl import load_workbook


def log_debug(msg):
    """输出调试信息"""
    print(f"[DEBUG] {msg}", file=sys.stderr)


def get_template_headers(template_path):
    """
    读取Excel模板的表头
    
    Args:
        template_path: Excel模板文件路径
        
    Returns:
        dict: {
            "success": True/False,
            "headers": 表头列表,
            "header_row": 表头行号,
            "message": 信息
        }
    """
    result = {
        "success": False,
        "headers": [],
        "header_row": None,
        "message": ""
    }
    
    try:
        log_debug(f"读取模板: {template_path}")
        
        # 检查文件是否存在
        if not os.path.exists(template_path):
            result["message"] = f"模板文件不存在: {template_path}"
            return result
        
        # 加载工作簿
        wb = load_workbook(template_path)
        sheet = wb.active
        
        # 查找表头行（包含常见字段的行）
        common_headers = ["姓名", "身份证号码", "手机联系方式", "邮箱", "性别"]
        header_row = None
        
        for row_idx in range(1, min(11, sheet.max_row + 1)):  # openpyxl使用1-based索引
            row_cells = list(sheet.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
            row_values = [str(cell) if cell is not None else "" for cell in row_cells]
            row_text = " ".join(row_values)
            
            # 检查是否包含常见表头
            matched_count = sum(1 for h in common_headers if h in row_text)
            if matched_count >= 2:  # 至少包含2个常见字段
                header_row = row_idx
                break
        
        if header_row is None:
            header_row = 1
        
        # 读取表头
        header_row_cells = list(sheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True))[0]
        headers = [str(cell) if cell is not None else "" for cell in header_row_cells]
        
        # 过滤空表头
        headers = [h for h in headers if h.strip()]
        
        result["success"] = True
        result["headers"] = headers
        result["header_row"] = header_row
        result["message"] = f"成功读取模板，表头包含{len(headers)}列"
        
        log_debug(f"表头行: {header_row}")
        log_debug(f"表头列数: {len(headers)}")
        
        return result
        
    except Exception as e:
        error_msg = f"读取模板失败: {str(e)}"
        log_debug(error_msg)
        log_debug(traceback.format_exc())
        result["message"] = error_msg
        return result


def main():
    """命令行入口"""
    try:
        if len(sys.argv) < 2:
            print("用法: python get_template_headers.py <template_path>")
            print("示例: python get_template_headers.py ./template.xlsx")
            sys.exit(1)
        
        template_path = sys.argv[1]
        result = get_template_headers(template_path)
        print(json.dumps(result, ensure_ascii=False, indent=2))
        sys.exit(0 if result["success"] else 1)
        
    except Exception as e:
        error_result = {
            "success": False,
            "headers": [],
            "header_row": None,
            "message": f"程序异常: {str(e)}"
        }
        print(json.dumps(error_result, ensure_ascii=False, indent=2))
        sys.exit(1)


if __name__ == "__main__":
    main()
